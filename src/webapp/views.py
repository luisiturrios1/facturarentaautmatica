# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.views import generic
from django.conf import settings
from django.contrib import messages
from django.urls import reverse_lazy
from django.forms import formset_factory
from django.shortcuts import get_object_or_404, render, HttpResponseRedirect, HttpResponse
from django.views.generic.edit import FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin

from dateutil.parser import parse as parse_date
from django_tables2 import SingleTableView
from django_tables2 import SingleTableMixin
from django_filters.views import FilterView
from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook
from cStringIO import StringIO
import xmltodict
import zipfile

from webapp import models
from webapp import tables
from webapp import filters
from webapp import forms
from webapp import tasks


class Index(LoginRequiredMixin, generic.TemplateView):
    template_name = 'webapp/index.html'

    def get_context_data(self, **kwargs):
        context = super(Index, self).get_context_data(**kwargs)
        context['razon_social'] = settings.RAZON_SOCIAL
        return context


class ContratoList(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = models.Contrato
    table_class = tables.ContratoTable
    filterset_class = filters.ContratoFilter
    paginate_by = 20
    template_name = 'webapp/contrato-list.html'


class ContratoExcel(LoginRequiredMixin, generic.View):
    def get(self, request, *args, **kwargs):
        contratos = models.Contrato.objects.all()

        wb = Workbook()

        ws = wb.active

        ws.append((
            'nombre_predio',
            'cuenta_predial',
            'nombre_cliente',
            'rfc_cliente',
            'correo_cliente',
            'telefono_cliente',
            'saldo_contrato',
            'precio_mensual',
            'retener_impuestos',
            'uso_cfdi',
            'metodo_pago',
            'forma_pago',
            'dia_facturacion',
            'fecha_registro',
            'fecha_modificado',
        ))

        for c in contratos:
            ws.append((
                c.nombre_predio,
                c.cuenta_predial,
                c.nombre_cliente,
                c.rfc_cliente,
                c.correo_cliente,
                c.telefono_cliente,
                c.saldo_contrato,
                c.precio_mensual,
                c.retener_impuestos,
                c.uso_cfdi,
                c.metodo_pago,
                c.forma_pago,
                c.dia_facturacion,
                c.fecha_registro,
                c.fecha_modificado
            ))

        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'filename="Contratos_{}.xlsx"'.format(settings.RFC)
        return response


class ContratoCreate(LoginRequiredMixin, SuccessMessageMixin, generic.CreateView):
    model = models.Contrato
    fields = [
        'nombre_predio',
        'cuenta_predial',
        'nombre_cliente',
        'rfc_cliente',
        'correo_cliente',
        'telefono_cliente',
        'saldo_contrato',
        'precio_mensual',
        'retener_impuestos',
        'uso_cfdi',
        'metodo_pago',
        'forma_pago',
        'dia_facturacion',
    ]
    success_message = u'Contrato %(nombre_predio)s creado con éxito'
    template_name = 'webapp/contrato-create.html'


class ContratoUpdate(LoginRequiredMixin, SuccessMessageMixin, generic.UpdateView):
    model = models.Contrato
    fields = [
        'nombre_predio',
        'cuenta_predial',
        'nombre_cliente',
        'rfc_cliente',
        'correo_cliente',
        'telefono_cliente',
        'saldo_contrato',
        'precio_mensual',
        'retener_impuestos',
        'uso_cfdi',
        'metodo_pago',
        'forma_pago',
        'dia_facturacion',
    ]
    success_message = u'Contrato %(nombre_predio)s modificado con éxito'
    template_name = 'webapp/contrato-update.html'


class ContratoDetail(LoginRequiredMixin, generic.DetailView):
    model = models.Contrato
    template_name = 'webapp/contrato-detail.html'


class ContratoDelete(LoginRequiredMixin, generic.DeleteView):
    model = models.Contrato
    success_message = u'Contrato borrado'
    success_url = reverse_lazy('webapp:contrato-list')
    template_name = 'webapp/contrato-delete.html'

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, self.success_message)
        return super(ContratoDelete, self).delete(request, *args, **kwargs)


class FacturaDetail(LoginRequiredMixin, generic.DetailView):
    model = models.Factura
    template_name = 'webapp/factura-detail.html'


class FacturaPagar(LoginRequiredMixin, FormView):
    template_name = 'webapp/factura-pagar.html'
    form_class = forms.FacturaPagarForm

    def get_success_url(self):
        return reverse_lazy('webapp:factura-detail', kwargs={'pk': self.kwargs['pk']})

    def get_context_data(self, **kwargs):
        context = super(FacturaPagar, self).get_context_data(**kwargs)
        context['factura'] = get_object_or_404(models.Factura, pk=self.kwargs['pk'])
        return context

    def form_valid(self, form):
        factura = get_object_or_404(models.Factura, pk=self.kwargs['pk'])
        contrato = factura.contrato

        factura.pagada = True
        factura.fecha_pago = form.cleaned_data['fecha_pago']
        factura.forma_pago_rep = form.cleaned_data['forma_pago_rep']
        factura.num_operacion = form.cleaned_data['num_operacion']
        factura.save()

        tasks.facturar_rep(factura.id)

        contrato.saldo_contrato = contrato.saldo_contrato - factura.total
        contrato.save()

        return super(FacturaPagar, self).form_valid(form)


class FacturaList(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = models.Factura
    table_class = tables.FacturaTable
    filterset_class = filters.FacturaFilter
    paginate_by = 20
    template_name = 'webapp/factura-list.html'


class FacturaZip(LoginRequiredMixin, generic.View):
    def get(self, request, *args, **kwargs):
        form = forms.FacturaExcelForm()
        return render(request, 'webapp/factura-zip.html', {'form': form})

    def post(self, request, *args, **kwargs):
        form = forms.FacturaExcelForm(request.POST)

        if form.is_valid():

            facturas = models.Factura.objects.filter(
                fecha__range=[form.cleaned_data['fecha_inicial'], form.cleaned_data['fecha_final']]
            )

            export_file = StringIO()

            with zipfile.ZipFile(export_file, 'w', zipfile.ZIP_DEFLATED) as writer:
                for f in facturas:
                    path = 'cfdis/{0}/{1}/{2}_{3}'.format(f.nombre_cliente, f.fecha.month, f.serie, f.folio)
                    writer.write(f.xml.path, path + '.xml')
                    writer.write(f.pdf.path, path + '.pdf')

            response = HttpResponse(
                export_file.getvalue(), content_type='application/x-zip-compressed'
            )

            response['Content-Disposition'] = 'attachment; filename="cfdis.zip"'

            return response


class FacturaExcel(LoginRequiredMixin, generic.View):
    def get(self, request, *args, **kwargs):
        form = forms.FacturaExcelForm()
        return render(request, 'webapp/factura-excel.html', {'form': form})

    def post(self, request, *args, **kwargs):

        form = forms.FacturaExcelForm(request.POST)

        if form.is_valid():

            facturas = models.Factura.objects.filter(
                fecha__range=[form.cleaned_data['fecha_inicial'], form.cleaned_data['fecha_final']]
            )

            wb = Workbook()

            ws = wb.active

            ws.append((
                'uuid',
                'serie',
                'folio',
                'fecha',
                'nombre_cliente',
                'rfc_cliente',
                'correo_cliente',
                'concepto',
                'sub_total',
                'iva_trasladado',
                'iva_retenido',
                'isr_retenido',
                'total',
                'contrato',
                'fecha_pago'
            ))

            for f in facturas:
                ws.append((
                    str(f.uuid),
                    f.serie,
                    f.folio,
                    f.fecha,
                    f.nombre_cliente,
                    f.rfc_cliente,
                    f.correo_cliente,
                    f.concepto,
                    f.sub_total,
                    f.iva_trasladado,
                    f.iva_retenido,
                    f.isr_retenido,
                    f.total,
                    f.contrato.nombre_predio,
                    f.fecha_pago
                ))

            response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'filename="Facturas_{}.xlsx"'.format(settings.RFC)
            return response
        else:
            return render(request, 'webapp/factura-excel.html', {'form': form})


class FacturaPdf(LoginRequiredMixin, generic.TemplateView):
    template_name = 'webapp/factura-pdf.html'

    def get_context_data(self, **kwargs):
        context = super(FacturaPdf, self).get_context_data(**kwargs)

        cfdi = get_object_or_404(models.Factura, pk=self.kwargs['pk'])

        context['xml'] = xmltodict.parse(cfdi.xml.read())['cfdi:Comprobante']

        return context


class ContratoFacturar(LoginRequiredMixin, FormView):
    template_name = 'webapp/contrato-facturar.html'
    form_class = forms.FacturarForm

    def get_context_data(self, **kwargs):
        context = super(ContratoFacturar, self).get_context_data(**kwargs)
        context['contrato'] = get_object_or_404(models.Contrato, pk=self.kwargs['pk'])
        return context

    def form_valid(self, form):
        contrato_id = self.kwargs['pk']

        descripcion = form.cleaned_data['concepto']
        sub_total = form.cleaned_data['sub_total']

        factura_id = tasks.facturar(contrato_id, descripcion=descripcion, sub_total=sub_total)

        self.success_url = reverse_lazy('webapp:factura-detail', kwargs={'pk': factura_id})

        return super(ContratoFacturar, self).form_valid(form)


class PagosBucar(LoginRequiredMixin, FormView):
    template_name = 'webapp/pagos-buscar.html'
    form_class = forms.PagoBuscarForm

    def form_valid(self, form):

        if 'estado_de_cuenta' in self.request.FILES:

            wb = load_workbook(self.request.FILES['estado_de_cuenta'])

            ws = wb.active

            result_list = []

            for i in range(5, ws.max_row + 1):
                t_list = [
                    ws['A{}'.format(i)].value,
                    ws['B{}'.format(i)].value,
                    ws['C{}'.format(i)].value,
                    ws['D{}'.format(i)].value,
                    ws['E{}'.format(i)].value,
                ]

                # Agregar solo los que son abono
                if t_list[0] is not None and t_list[2] is None:
                    t_list[0] = t_list[0].isoformat()
                    result_list.append(t_list)

            self.request.session['estado_de_cuenta'] = result_list

        else:
            if 'estado_de_cuenta' in self.request.session:
                del self.request.session['estado_de_cuenta']

        self.success_url = reverse_lazy(
            'webapp:pagos-consolidar',
            kwargs={
                'fecha_inicial': form.cleaned_data['fecha_inicial'].strftime('%Y-%m-%d'),
                'fecha_final': form.cleaned_data['fecha_final'].strftime('%Y-%m-%d'),
            }
        )

        return super(PagosBucar, self).form_valid(form)


class PagosConsolidar(LoginRequiredMixin, generic.View):
    template_name = 'webapp/pagos-consolidar.html'
    form_class = forms.PagoConsolidarForm

    def get(self, request, *args, **kwargs):

        fecha_inicial = self.kwargs.get('fecha_inicial')
        fecha_final = self.kwargs.get('fecha_final')

        FormSet = formset_factory(self.form_class, extra=0)

        facturas = models.Factura.objects.filter(pagada=False, fecha__range=[fecha_inicial, fecha_final])

        if 'estado_de_cuenta' in self.request.session:
            estado_de_cuenta = self.request.session['estado_de_cuenta']
        else:
            estado_de_cuenta = []

        initial = []

        for factura in facturas:
            tmp = {
                'id_factura': factura.id,
                'fecha_factura': factura.fecha,
                'uuid': factura.uuid,
                'folio': '{} {}'.format(factura.serie, factura.folio),
                'nombre_cliente': factura.nombre_cliente,
                'rfc_cliente': factura.rfc_cliente,
                'concepto': factura.concepto,
                'total': factura.total,
                'moviento_cuenta': '',
            }

            for mvt in estado_de_cuenta:
                if int(mvt[3]) == int(factura.total):
                    tmp['fecha_pago'] = parse_date(mvt[0]).strftime('%Y-%m-%d')
                    tmp['moviento_cuenta'] = '{0}\n{1}\n{2}'.format(mvt[0], mvt[1], mvt[3])
                    break

            initial.append(tmp)

        formset = FormSet(initial=initial)

        return render(
            request, self.template_name,
            {'fecha_inicial': fecha_inicial, 'fecha_final': fecha_final, 'formset': formset}
        )

    def post(self, request, *args, **kwargs):

        fecha_inicial = self.kwargs.get('fecha_inicial')
        fecha_final = self.kwargs.get('fecha_final')

        FormSet = formset_factory(self.form_class, extra=0)

        facturas = models.Factura.objects.filter(pagada=False, fecha__range=[fecha_inicial, fecha_final])

        if 'estado_de_cuenta' in self.request.session:
            estado_de_cuenta = self.request.session['estado_de_cuenta']
        else:
            estado_de_cuenta = []

        initial = []

        for factura in facturas:
            tmp = {
                'id_factura': factura.id,
                'fecha_factura': factura.fecha,
                'uuid': factura.uuid,
                'folio': '{} {}'.format(factura.serie, factura.folio),
                'nombre_cliente': factura.nombre_cliente,
                'rfc_cliente': factura.rfc_cliente,
                'concepto': factura.concepto,
                'total': factura.total,
                'moviento_cuenta': '',
            }

            for mvt in estado_de_cuenta:
                if int(mvt[3]) == int(factura.total):
                    tmp['fecha_pago'] = parse_date(mvt[0]).strftime('%Y-%m-%d')
                    tmp['moviento_cuenta'] = '{0}\n{1}\n{2}'.format(mvt[0], mvt[1], mvt[3])
                    break

            initial.append(tmp)

        formset = FormSet(request.POST, initial=initial)

        if formset.is_valid():
            for form in formset:
                id_factura = form.cleaned_data['id_factura']
                fecha_pago = form.cleaned_data['fecha_pago']

                if fecha_pago is not None:
                    f = models.Factura.objects.get(id=id_factura)
                    f.fecha_pago = fecha_pago
                    f.pagada = True
                    f.save()

                    c = f.contrato
                    c.saldo_contrato = c.saldo_contrato - f.total
                    c.save()

            messages.success(request, 'Facturas pagadas con exito')
            return HttpResponseRedirect(reverse_lazy('webapp:factura-list'))

        return render(
            request, self.template_name,
            {'fecha_inicial': fecha_inicial, 'fecha_final': fecha_final, 'formset': formset}
        )
